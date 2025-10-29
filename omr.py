import os
from PIL import Image, ImageOps
import pandas as pd
import glob
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import numpy as np
import cv2
import pytesseract
from pdf2image import convert_from_path

load_dotenv(override=True)

def align_image(gray_image):
    """Alinha a imagem corrigindo a inclinação detectada"""
    gray_cv = np.array(gray_image)

    if gray_cv.dtype != np.uint8:
        gray_cv = gray_cv.astype(np.uint8)

    # Detecção de bordas
    edges = cv2.Canny(gray_cv, 50, 150, apertureSize=3)

    # Encontrar linhas usando Hough Transform
    lines = cv2.HoughLinesP(
        edges,
        rho=1,
        theta=np.pi / 180,
        threshold=100,
        minLineLength=100,
        maxLineGap=10
    )

    if lines is None or len(lines) == 0:
        print("Nenhuma linha detectada, mantendo imagem original.")
        return gray_image

    # Calcular ângulos das linhas detectadas
    angles = []
    for line in lines:
        x1, y1, x2, y2 = line[0]
        angle_rad = np.arctan2(y2 - y1, x2 - x1)
        angle_deg = np.degrees(angle_rad)
        
        # Normalizar ângulo para o intervalo [-90, 90]
        if angle_deg < -90:
            angle_deg += 180
        elif angle_deg > 90:
            angle_deg -= 180
        angles.append(angle_deg)

    # Calcular ângulo mediano
    median_angle = np.median(angles)

    if abs(median_angle) < 0.1:
        print("Imagem já está alinhada.")
        return gray_image

    # Rotacionar imagem para corrigir inclinação
    (h, w) = gray_cv.shape
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, median_angle, 1.0)
    aligned_gray_cv = cv2.warpAffine(
        gray_cv,
        M,
        (w, h),
        flags=cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_REPLICATE
    )

    aligned_gray_image = Image.fromarray(aligned_gray_cv)
    print(f"Imagem rotacionada em {median_angle:.2f} graus")
    
    return aligned_gray_image


def find_reference_rectangle(aligned_image, debug_mode=False):
    """
    Encontra o primeiro retângulo de referência na imagem (campo "Nome do Candidato").
    Usado apenas para extrair o identificador do exame.
    """
    image_cv = np.array(aligned_image)

    if len(image_cv.shape) == 3:
        image_cv = cv2.cvtColor(image_cv, cv2.COLOR_RGB2GRAY)

    # Aplicar threshold binário
    _, binary_image = cv2.threshold(image_cv, 128, 255, cv2.THRESH_BINARY_INV)

    # Operações morfológicas para limpar ruído
    kernel = np.ones((5, 5), np.uint8)
    binary_image = cv2.morphologyEx(binary_image, cv2.MORPH_CLOSE, kernel)

    # Encontrar contornos
    contours, _ = cv2.findContours(
        binary_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
    )

    rectangles = []

    # Parâmetros configuráveis
    min_area = int(os.getenv("reference_rect_min_area", 10000))
    max_area = int(os.getenv("reference_rect_max_area", 200000))
    min_aspect_ratio = float(os.getenv("reference_rect_min_aspect_ratio", 3.0))
    max_aspect_ratio = float(os.getenv("reference_rect_max_aspect_ratio", 12.0))
    min_extent = float(os.getenv("reference_rect_min_extent", 0.6))
    max_y_position = int(os.getenv("reference_rect_max_y_position", 400))

    for contour in contours:
        area = cv2.contourArea(contour)

        if area < min_area or area > max_area:
            continue

        x, y, w, h = cv2.boundingRect(contour)

        if y > max_y_position:
            continue
        
        if h == 0:
            continue
            
        aspect_ratio = float(w) / h
        
        if aspect_ratio < min_aspect_ratio or aspect_ratio > max_aspect_ratio:
            continue

        rect_area = w * h
        if rect_area == 0:
            continue
            
        extent = float(area) / rect_area
        
        if extent < min_extent:
            continue

        rectangles.append({
            'x': x,
            'y': y,
            'w': w,
            'h': h,
            'area': area,
            'aspect_ratio': aspect_ratio,
            'extent': extent
        })

    if len(rectangles) == 0:
        print("⚠️  AVISO: Nenhum retângulo de referência encontrado!")
        return None

    # Ordenação inteligente: priorizar mais à esquerda quando Y é similar
    y_threshold = int(os.getenv("reference_rect_y_similarity_threshold", 30))
    
    rectangles = sorted(rectangles, key=lambda r: r['x'])
    
    if len(rectangles) > 1:
        min_y = min(r['y'] for r in rectangles)
        top_rectangles = [r for r in rectangles if r['y'] <= min_y + y_threshold]
        if top_rectangles:
            rectangles = sorted(top_rectangles, key=lambda r: r['x']) + \
                        [r for r in rectangles if r not in top_rectangles]

    first_rect = rectangles[0]
    
    print(f"\n✅ Retângulo de referência encontrado: x={first_rect['x']}, y={first_rect['y']}")

    if debug_mode:
        for idx, rect in enumerate(rectangles[:3]):
            x, y, w, h = rect['x'], rect['y'], rect['w'], rect['h']
            rect_img = aligned_image.crop((x, y, x + w, y + h))
            fname = f"debug_rectangle_{idx+1}.png"
            rect_img.save(fname)
            print(f"  🔍 Debug: Retângulo #{idx+1} salvo em '{fname}'")
    
    return first_rect


def extract_exam_identifier(aligned_image, reference_rect, debug_mode=False):
    """
    Extrai o identificador do exame usando OCR.
    Se não houver retângulo de referência, tenta buscar em posição absoluta.
    """
    if reference_rect is None:
        # Fallback: posição absoluta configurável
        print("⚠️  Usando posição absoluta para identificador (sem retângulo de referência)")
        id_x = int(os.getenv("exam_id_absolute_x", 1000))
        id_y = int(os.getenv("exam_id_absolute_y", 200))
        id_width = int(os.getenv("exam_identifier_width", 300))
        id_height = int(os.getenv("exam_identifier_height", 80))
    else:
        # Posição relativa ao retângulo de referência
        id_offset_x = int(os.getenv("exam_id_offset_from_ref_x", 800))
        id_offset_y = int(os.getenv("exam_id_offset_from_ref_y", 0))
        id_width = int(os.getenv("exam_identifier_width", 300))
        id_height = int(os.getenv("exam_identifier_height", 80))

        id_x = reference_rect['x'] + id_offset_x
        id_y = reference_rect['y'] + id_offset_y

    id_x = max(0, id_x)
    id_y = max(0, id_y)

    id_area = aligned_image.crop((id_x, id_y, id_x + id_width, id_y + id_height))

    if debug_mode:
        id_area.save('debug_exam_identifier_area.png')
        print(f"🔍 Debug: Área do identificador salva em 'debug_exam_identifier_area.png'")

    exam_identifier = pytesseract.image_to_string(id_area).strip()

    print(f"Identificador extraído: '{exam_identifier}'")
    
    return exam_identifier if exam_identifier else "ID_NAO_IDENTIFICADO"


# Configurações principais
base_path = os.getenv("base_path", ".")
images_folder_path = f"{base_path}/arquivos_escaneados"
template_path = f"{base_path}/gabarito.xlsx"

debug_mode = os.getenv("DEBUG_MODE", "false").lower() == "true"

images_list = glob.glob(f"{images_folder_path}/*")

if not images_list:
    print(f"⚠️  AVISO: Nenhuma imagem encontrada em '{images_folder_path}'")
    exit(1)

results = []
df = pd.read_excel(template_path)
template_dict = df.set_index('Questão')['Resposta'].to_dict()

# Parâmetros das questões
question_blocks_quantity = int(os.getenv("question_blocks_quantity"))
questions_per_block_quantity = int(os.getenv("questions_per_block_quantity"))
total_questions = int(os.getenv("total_questions"))
question_index_initial_value = int(os.getenv("question_index_initial_value"))
dark_pixel_threshold = int(os.getenv("dark_pixel_threshold"))

# Adicionar gabarito aos resultados
results.append({
    'exam_identifier': 'Gabarito',
    'file': 'N/A',
    'correct_questions_quantity': total_questions,
    'correct_questions_percentage': 100,
    'answers': template_dict
})

# Processar cada arquivo
for file_path in images_list:
    try:
        print(f"\n{'='*80}")
        print(f"PROCESSANDO: {file_path}")
        print(f"{'='*80}")

        # Carregar imagem
        if file_path.lower().endswith('.pdf'):
            pages = convert_from_path(file_path)
            image = pages[0]
        else:
            image = Image.open(file_path)

        # Converter para escala de cinza
        gray_image = ImageOps.grayscale(image)

        # Alinhar imagem
        aligned_gray_image = align_image(gray_image)

        # PASSO 1: Encontrar retângulo de referência (apenas para ID)
        reference_rect = find_reference_rectangle(aligned_gray_image, debug_mode=debug_mode)

        # PASSO 2: Extrair identificador do exame
        exam_identifier = extract_exam_identifier(aligned_gray_image, reference_rect, debug_mode=debug_mode)

        # PASSO 3: Definir primeira bolha a partir do retângulo de referência
        column_start_x, row_start_y = reference_rect['x'], reference_rect['y']

        if column_start_x is None or row_start_y is None:
            print(f"❌ ERRO: Não foi possível encontrar bolhas em {file_path}")
            print("⏭️  Pulando este arquivo...\n")
            continue
        
        column_start_x += int(os.getenv("first_bubble_offset_from_ref_x", 100))
        row_start_y += int(os.getenv("first_bubble_offset_from_ref_y", 200))

        # PASSO 4: Processar respostas
        binary_image = aligned_gray_image.point(lambda x: 0 if x < 128 else 255, '1')

        alternatives_quantity = 5
        alternatives = ['A', 'B', 'C', 'D', 'E']

        column_spacing = int(os.getenv("column_spacing"))
        block_spacing = int(os.getenv("block_spacing"))
        row_spacing = int(os.getenv("row_spacing"))

        circle_width = int(os.getenv("circle_width"))
        circle_height = int(os.getenv("circle_height"))

        student_answers = {}
        question_index = question_index_initial_value - 1
        correct_questions_quantity = 0

        for block_index in range(question_blocks_quantity):
            for row_index in range(questions_per_block_quantity):
                row = row_start_y + row_index * row_spacing + row_index * circle_height
                
                for col_index in range(alternatives_quantity):
                    col = (column_start_x + 
                           col_index * column_spacing + 
                           col_index * circle_width + 
                           block_index * block_spacing + 
                           block_index * alternatives_quantity * circle_width + 
                           block_index * alternatives_quantity * column_spacing)
                    
                    mark_area = binary_image.crop((
                        col, 
                        row, 
                        col + circle_width, 
                        row + circle_height
                    ))

                    if debug_mode:
                        question_area = aligned_gray_image.crop((col, row, col + circle_width, row + circle_height))
                        question_area.save(f"debug_question_{question_index+1}_item_{col_index}.png")
                    
                    black_pixels_count = sum(1 for pixel in mark_area.getdata() if pixel == 0)
                    
                    if black_pixels_count > dark_pixel_threshold:
                        student_answers[question_index + 1] = alternatives[col_index]
                
                if not student_answers.get(question_index + 1):
                    student_answers[question_index + 1] = '-'
                
                if student_answers.get(question_index + 1) == template_dict.get(question_index + 1):
                    correct_questions_quantity += 1
                
                question_index += 1
                
                if question_index + 1 > question_index_initial_value + total_questions:
                    break
            
            if question_index + 1 > question_index_initial_value + total_questions:
                break

        results.append({
            'exam_identifier': exam_identifier,
            'file': Path(file_path).stem,
            'correct_questions_quantity': correct_questions_quantity,
            'correct_questions_percentage': round((correct_questions_quantity / total_questions) * 100, 2),
            'answers': student_answers
        })

        print(f"\n✅ Processamento concluído: {correct_questions_quantity}/{total_questions} corretas "
              f"({round((correct_questions_quantity / total_questions) * 100, 2)}%)")

        # Se estiver em modo debug, parar após primeiro arquivo
        if debug_mode:
            print("\n" + "="*80)
            print("🔍 MODO DEBUG - Processamento interrompido após primeiro arquivo")
            print("="*80)
            print("\nArquivos de debug gerados:")
            print("  - debug_circle_detection.png: Círculos detectados")
            print("  - debug_rectangle_*.png: Retângulos detectados (se houver)")
            print("  - debug_exam_identifier_area.png: Área do identificador")
            print("\nAjuste os parâmetros no .env e teste novamente.")
            print("Para processar todos: DEBUG_MODE=false\n")
            exit(0)

    except Exception as e:
        print(f"\n❌ ERRO ao processar {file_path}: {str(e)}")
        import traceback
        traceback.print_exc()
        continue

# Gerar planilha de resultados
print(f"\n{'='*80}")
print("📊 GERANDO PLANILHA DE RESULTADOS...")
print(f"{'='*80}")

file_name = base_path + '/resultados_' + datetime.now().strftime("%Y%m%d-%H%M%S") + '.xlsx'
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Gabarito e resultados"

spreadsheet_row_index = 1
worksheet.cell(row=spreadsheet_row_index, column=1).value = 'Arquivo'
worksheet.column_dimensions['A'].width = 50
worksheet.cell(row=spreadsheet_row_index, column=2).value = 'Identificador da prova'
worksheet.column_dimensions['B'].width = 50
worksheet.cell(row=spreadsheet_row_index, column=3).value = 'Questões corretas'
worksheet.column_dimensions['C'].width = 20
worksheet.cell(row=spreadsheet_row_index, column=4).value = 'Percentual de acerto'
worksheet.column_dimensions['D'].width = 20

for column_index in range(total_questions):
    worksheet.column_dimensions[get_column_letter(column_index + 5)].width = 5
    worksheet.cell(row=spreadsheet_row_index, column=column_index + 5).value = (
        question_index_initial_value + column_index
    )

spreadsheet_row_index += 1
for result in results:
    worksheet.cell(row=spreadsheet_row_index, column=1).value = result['file']
    worksheet.cell(row=spreadsheet_row_index, column=2).value = result['exam_identifier']
    worksheet.cell(row=spreadsheet_row_index, column=3).value = result['correct_questions_quantity']
    worksheet.cell(row=spreadsheet_row_index, column=4).value = result['correct_questions_percentage']
    
    for column_index in range(total_questions):
        worksheet.cell(row=spreadsheet_row_index, column=column_index + 5).value = (
            result['answers'][question_index_initial_value + column_index]
        )
    
    spreadsheet_row_index += 1

workbook.save(file_name)
workbook.close()

print(f"\n✅ Planilha salva em: {file_name}")
print(f"✅ Total de arquivos processados: {len(results) - 1}")
print(f"{'='*80}\n")