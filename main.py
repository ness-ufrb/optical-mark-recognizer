import os
from PIL import Image, ImageOps
import pandas as pd
import glob
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import numpy as np
import cv2
import pytesseract
from pdf2image import convert_from_path
# import easyocr

load_dotenv(override=True)

def align_image(gray_image):
    # Convert PIL grayscale image to OpenCV format (NumPy array)
    gray_cv = np.array(gray_image)

    # Check if the image is already in uint8 format
    if gray_cv.dtype != np.uint8:
        gray_cv = gray_cv.astype(np.uint8)

    # Use edge detection (Canny Edge Detection)
    edges = cv2.Canny(gray_cv, 50, 150, apertureSize=3)

    # Find lines using Probabilistic Hough Line Transform
    lines = cv2.HoughLinesP(
        edges,
        rho=1,
        theta=np.pi / 180,
        threshold=100,
        minLineLength=100,
        maxLineGap=10
    )

    # If no lines are found, return the original image
    if lines is None or len(lines) == 0:
        print("No lines detected, unable to align image.")
        return gray_image

    # Calculate the angles of the detected lines
    angles = []
    for line in lines:
        x1, y1, x2, y2 = line[0]
        angle_rad = np.arctan2(y2 - y1, x2 - x1)
        angle_deg = np.degrees(angle_rad)
        # Normalize angle to the range [-90, 90]
        if angle_deg < -90:
            angle_deg += 180
        elif angle_deg > 90:
            angle_deg -= 180
        angles.append(angle_deg)

    # Compute the median angle of the detected lines
    median_angle = np.median(angles)

    # If the median angle is near zero, the image is already aligned
    if abs(median_angle) < 0.1:
        print("Image is already aligned.")
        return gray_image

    # Rotate image to correct skew
    (h, w) = gray_cv.shape
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, median_angle, 1.0)
    aligned_gray_cv = cv2.warpAffine(
        gray_cv,
        M,
        (w, h),
        flags=cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_REPLICATE
    )

    # Convert back to PIL Image format
    aligned_gray_image = Image.fromarray(aligned_gray_cv)

    return aligned_gray_image


def find_first_bubble_position(aligned_image):
    # Convert PIL image to OpenCV format
    image_cv = np.array(aligned_image)

    # Check if the image has multiple channels (e.g., RGB)
    if len(image_cv.shape) == 3:
        # If so, convert to grayscale
        image_cv = cv2.cvtColor(image_cv, cv2.COLOR_RGB2GRAY)

    # Apply binary thresholding (invert image so bubbles are white on black background)
    _, binary_image = cv2.threshold(image_cv, 128, 255, cv2.THRESH_BINARY_INV)

    # Optionally, apply morphological operations to remove small noise
    kernel = np.ones((3, 3), np.uint8)
    binary_image = cv2.morphologyEx(binary_image, cv2.MORPH_OPEN, kernel)

    # Find contours in the binary image
    contours, _ = cv2.findContours(
        binary_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
    )

    # List to hold bubble contours and their bounding rectangles
    bubbles = []

    # Get minimum bubble size from environment variable or set default
    min_bubble_size = int(os.getenv("min_bubble_size", 20))

    # Loop over the contours
    for contour in contours:
        # Calculate area
        area = cv2.contourArea(contour)

        # Filter contours based on area (adjust thresholds as needed)
        if area < 100 or area > 10000:
            continue  # Skip contours that are too small or too large

        # Get bounding rectangle
        x, y, w, h = cv2.boundingRect(contour)

        # Filter based on width and height (minimum size)
        if w < min_bubble_size or h < min_bubble_size:
            continue  # Skip contours smaller than minimum size

        # Optionally, filter based on aspect ratio (if bubbles are approximately circular)
        aspect_ratio = float(w) / h
        if aspect_ratio < 0.8 or aspect_ratio > 1.2:
            continue  # Skip contours that are not roughly square

        # Calculate perimeter
        perimeter = cv2.arcLength(contour, True)

        # Calculate circularity
        circularity = 4 * np.pi * (area / (perimeter * perimeter))

        # Filter based on circularity
        if circularity < 0.7:
            continue  # Skip non-circular contours

        # Append to bubbles list
        bubbles.append({'contour': contour, 'rect': (x, y, w, h)})

    if len(bubbles) < 1:
        print("No bubbles detected in the image.")
        return None, None

    # Sort the bubbles from top to bottom, then left to right, using 'rect' positions
    bubbles = sorted(bubbles, key=lambda b: (b['rect'][1], b['rect'][0]))

    # Get the position of the first bubble
    first_bubble = bubbles[0]
    x, y, w, h = first_bubble['rect']
    # Return the initial x and y position (top-left corner)
    calculated_x = x - int(os.getenv("second_bubble_x_offset"))
    calculated_y = y + int(os.getenv("second_bubble_y_offset"))
    return calculated_x, calculated_y

# EASYOCR APPROACH
# ocr_reader = easyocr.Reader(['pt'])

base_path = os.getenv("base_path")
images_folder_path = f"{base_path}/imagens"
template_path = f"{base_path}/gabarito.xlsx"

images_list = glob.glob(f"{images_folder_path}/*")

results = []
df = pd.read_excel(template_path)
template_dict = df.set_index('Questão')['Resposta'].to_dict()

# Quantity of questions
question_blocks_quantity = int(os.getenv("question_blocks_quantity"))
questions_per_block_quantity = int(os.getenv("questions_per_block_quantity"))
total_questions = int(os.getenv("total_questions"))
question_index_initial_value = int(os.getenv("question_index_initial_value"))
black_pixels_threshold = int(os.getenv("black_pixels_threshold"))

results.append({
    'exam_identifier': 'Gabarito',
    'file': 'N/A',
    'correct_questions_quantity': total_questions,
    'correct_questions_percentage': 100,
    'answers': template_dict
})

for file_path in images_list:
    # Check if file is a PDF
    if file_path.lower().endswith('.pdf'):
        # Convert PDF to images
        pages = convert_from_path(file_path)
        image = pages[0]  # Assuming the first page is the one we need
    else:
        # Load the image directly if it's not a PDF
        image = Image.open(file_path)

    print(f"Processing {file_path}")

    # Converting the Image to Grayscale
    gray_image = ImageOps.grayscale(image)

    # Align the image
    aligned_gray_image = align_image(gray_image)

    # Getting exam_identifier
    exam_identifier_start_x = int(os.getenv("exam_identifier_start_x"))
    exam_identifier_start_y = int(os.getenv("exam_identifier_start_y"))
    exam_identifier_height = int(os.getenv("exam_identifier_height"))
    exam_identifier_width = int(os.getenv("exam_identifier_width"))
    exam_identifier_area = aligned_gray_image.crop((exam_identifier_start_x, exam_identifier_start_y, exam_identifier_start_x + exam_identifier_width, exam_identifier_start_y + exam_identifier_height))

    #### FOR TESTING PURPOSES ONLY ####
    # exam_identifier_area.save('exam_identifier_area.png')
    # exit()
    #### FOR TESTING PURPOSES ONLY ####

    # Specify the whitelist of characters (numbers, letters, and space)
    custom_config = "-c tessedit_char_whitelist=" + os.getenv("exam_identifier_whitelist") + ' --psm 6'

    # Use pytesseract with the custom configuration
    exam_identifier = pytesseract.image_to_string(exam_identifier_area, config=custom_config)

    # EASYOCR APPROACH
    # img_byte_arr = io.BytesIO()
    # exam_identifier_area.save(img_byte_arr, format='PNG')
    # img_byte_arr = img_byte_arr.getvalue()
    # ocr_result = ocr_reader.readtext(img_byte_arr)
    
    # exam_identifier = ''
    # for result_set in ocr_result:
    #     for item in result_set:
    #         if (isinstance(item, str)) and len(item) > 5:
    #             exam_identifier = item

    # Black pixel limit verification
    dark_pixel_threshold = int(os.getenv("dark_pixel_threshold"))

    # Convert the image to binary format
    binary_image = aligned_gray_image.point(lambda x: 0 if x < 128 else 255, '1')

    # Array to store the student number
    student_number = []

    # Quantity of alternatives
    alternatives_quantity = 5
    alternatives = ['A', 'B', 'C', 'D', 'E']

    # x, y coordinates of the first bubble
    column_start_x, row_start_y = find_first_bubble_position(aligned_gray_image)

    # The spacing between columns
    column_spacing = int(os.getenv("column_spacing"))
    block_spacing = int(os.getenv("block_spacing"))

    # The spacing between rows
    row_spacing = int(os.getenv("row_spacing"))

    # The width and height of each circle
    circle_width = int(os.getenv("circle_width"))
    circle_height = int(os.getenv("circle_height"))

    # Array to store the student number
    student_answers = {}
    question_index = question_index_initial_value - 1
    correct_questions_quantity = 0

    # Check the density of the marked area for each column and row
    for block_index in range(question_blocks_quantity):  # For 4 blocks
        for row_index in range(questions_per_block_quantity):  # For 30 questions
            row = row_start_y + row_index * row_spacing + row_index * circle_height
            for col_index in range(alternatives_quantity):  # For 5 columns
                col = column_start_x + \
                    col_index * column_spacing + \
                    col_index * circle_width + \
                    block_index * block_spacing + \
                    block_index * alternatives_quantity * circle_width + \
                    block_index * (alternatives_quantity - 1) * column_spacing
                # Crop the relevant area
                mark_area = binary_image.crop((col, row, col + circle_width, row + circle_height))
                
                #### FOR TESTING PURPOSES ONLY ####
                # if question_index + 1 == 1 or question_index + 1 == 2 or question_index + 1 == 10  or question_index + 1 == 11:
                #     question_area = aligned_gray_image.crop((col, row, col + circle_width, row + circle_height))
                #     question_area.save(f"question_{question_index+1}_item_{col_index}.png")
                #### FOR TESTING PURPOSES ONLY ####
                
                # Count the number of black pixels
                black_pixels_count = sum(1 for pixel in mark_area.getdata() if pixel == 0)
                # If there are more black pixels than the threshold value, accept it as marked
                if black_pixels_count > black_pixels_threshold:
                    student_answers[question_index + 1] = alternatives[col_index]
            if not student_answers.get(question_index + 1):
                student_answers[question_index + 1] = '-'
            if student_answers.get(question_index + 1) == template_dict.get(question_index + 1):
                correct_questions_quantity += 1
            question_index += 1
            if question_index + 1 > question_index_initial_value + total_questions:
                break
        if question_index + 1 > question_index_initial_value + total_questions:
            break
    #### FOR TESTING PURPOSES ONLY ####
    # exit()
    #### FOR TESTING PURPOSES ONLY ####

    results.append({
        'exam_identifier': exam_identifier,
        'file': Path(file_path).stem,
        'correct_questions_quantity': correct_questions_quantity,
        'correct_questions_percentage': (correct_questions_quantity / total_questions) * 100,
        'answers': student_answers
    })

file_name = base_path + '/resultados_' + datetime.now().strftime("%Y%m%d-%H%M%S") + '.xlsx'
workbook = Workbook()

worksheet = workbook.active
worksheet.title = "Gabarito e resultados"

spreadsheet_row_index = 1
worksheet.cell(row=spreadsheet_row_index, column=1).value = 'Arquivo'
worksheet.column_dimensions['A'].width = 50
worksheet.cell(row=spreadsheet_row_index, column=2).value = 'Identificador da prova'
worksheet.column_dimensions['B'].width = 50
worksheet.cell(row=spreadsheet_row_index, column=3).value = 'Questões corretas'
worksheet.column_dimensions['C'].width = 20
worksheet.cell(row=spreadsheet_row_index, column=4).value = 'Percentual de acerto'
worksheet.column_dimensions['D'].width = 20
for column_index in range(total_questions):
    worksheet.column_dimensions[get_column_letter(column_index + 5)].width = 5
    worksheet.cell(row=spreadsheet_row_index, column=column_index + 5).value = question_index_initial_value + column_index
spreadsheet_row_index += 1
for result in results:
    worksheet.cell(row=spreadsheet_row_index, column=1).value = result['file']
    worksheet.cell(row=spreadsheet_row_index, column=2).value = result['exam_identifier']
    worksheet.cell(row=spreadsheet_row_index, column=3).value = result['correct_questions_quantity']
    worksheet.cell(row=spreadsheet_row_index, column=4).value = result['correct_questions_percentage']
    for column_index in range(total_questions):
        worksheet.cell(row=spreadsheet_row_index, column=column_index + 5).value = result['answers'][question_index_initial_value + column_index]
    spreadsheet_row_index += 1

workbook.save(file_name)
workbook.close()
